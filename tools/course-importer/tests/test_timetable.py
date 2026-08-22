from __future__ import annotations

import io
import json
import logging
import shutil
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from jsonschema import Draft202012Validator
from openpyxl import Workbook
from openpyxl.styles import Border, Side


TOOL_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = TOOL_ROOT.parents[1]
sys.path.insert(0, str(TOOL_ROOT))

from app import create_app  # noqa: E402
from importer import ImporterError  # noqa: E402
from timetable_service import DEFAULT_PERIODS, TimetableError, TimetableService  # noqa: E402


class TimetableServiceTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.repo = root / "repo"
        self.work = root / "work"
        (self.repo / "data").mkdir(parents=True)
        for name in (
            "timetable-month.schema.json",
            "timetable-manifest.schema.json",
            "timetable-periods.schema.json",
        ):
            shutil.copy2(REPO_ROOT / "data" / name, self.repo / "data" / name)
        (self.repo / "data" / "courses.json").write_text(
            json.dumps(
                {
                    "courses": [
                        {"id": "web-programming", "title": "Webプログラミング"},
                        {"id": "java", "title": "Java"},
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        self.service = TimetableService(self.repo, self.work)
        self.service.runtime_root.mkdir(parents=True, exist_ok=True)
        self.service.config_path.write_text(
            json.dumps(
                {
                    "sheetName": "2025通年時間割",
                    "subjectMappings": {
                        "Webプログラミング": "web-programming",
                        "Webプログラミング[A]": "web-programming",
                        "Java": "java",
                        "Java[A]": "java",
                        "Java[C]": "java",
                        "Java[A][C]": "java",
                    },
                    "canonicalGroups": {
                        "group-a": {"displayName": "A組"},
                        "group-1": {"displayName": "1組"},
                    },
                    "groupMappings": {"(A組)": "group-a", "(1組)": "group-1"},
                    "legacyGroupTags": {"A": "選択A", "C": "選択C"},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    def tearDown(self):
        logger = logging.getLogger(f"classview-admin-{self.work.resolve()}")
        for handler in list(logger.handlers):
            handler.close()
            logger.removeHandler(handler)
        self.temp.cleanup()

    @staticmethod
    def workbook_bytes(*, unknown_subject: bool = False, suspicious_room: bool = False) -> io.BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2025通年時間割"
        sheet["B1"] = date(2025, 4, 21)
        for offset, column in enumerate(("E", "H", "K", "N", "Q", "T"), start=1):
            sheet[f"{column}1"] = f"=B1+{offset}"

        sheet.merge_cells("A2:A3")
        sheet["A2"] = "1限\n09:20–10:50"
        sheet["B2"] = "Webプログラミング（A組）【CDP】"
        sheet["B3"] = "Java[A][C]【山田・佐藤】" if not unknown_subject else "AI応用演習【専任】"
        sheet.merge_cells("C2:C3")
        sheet["C2"] = "1"
        sheet.merge_cells("D2:D3")
        sheet["D2"] = 721.722 if suspicious_room else 721722
        sheet["D2"].number_format = "000,000"

        sheet["A4"] = "2限 11:00-12:30"
        sheet["B4"] = "試験：Webプログラミング[予定]"
        sheet["C4"] = "2,3"
        sheet["D4"] = "733,734"

        sheet["A5"] = "3限 13:30-15:00"
        sheet["B5"] = "企業説明会"
        sheet["C5"] = "全学年"
        sheet["D5"] = "アリーナ"

        sheet["A6"] = "4限 15:10-16:40"
        sheet["B6"] = "冬期休暇"
        sheet["C6"] = "全学年"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def analyze(self, **kwargs):
        return self.service.analyze(
            self.workbook_bytes(**kwargs),
            "2025通年時間割.xlsx",
            "2025-04-21",
            "2025-04-27",
            sheet_name="2025通年時間割",
        )

    @staticmethod
    def classified_workbook_bytes() -> io.BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2025通年時間割"
        sheet["B1"] = date(2025, 4, 21)
        for offset, column in enumerate(("E", "H", "K", "N", "Q", "T"), start=1):
            sheet[f"{column}1"] = f"=B1+{offset}"
        for column in ("B", "E", "H", "K", "N", "Q", "T"):
            sheet[f"{column}2"] = "科目【講師】"
            sheet.cell(row=2, column=sheet[column + "2"].column + 1, value="学年")
            sheet.cell(row=2, column=sheet[column + "2"].column + 2, value="教室")

        bottom = Border(bottom=Side(style="thin"))
        sheet["A3"] = "1限 09:20-10:50"
        sheet["B3"], sheet["C3"], sheet["D3"] = "Java[A]【講師A】", "1", "101"
        sheet["B4"], sheet["C4"], sheet["D4"] = "Java[C]【講師B】", "1", "102"
        sheet["A5"] = "～"
        sheet["E5"], sheet["F5"], sheet["G5"] = "Java[Z]【講師A】", "2", "103"
        sheet["A5"].border = bottom

        sheet["A6"] = "2限 11:00-12:30"
        sheet["B6"], sheet["C6"], sheet["D6"] = "Java[A]【講師A】", "1", "101"
        sheet["B7"], sheet["C7"], sheet["D7"] = "Java[A]【講師A】", "2", "101"
        sheet["B8"], sheet["C8"], sheet["D8"] = "Webプログラミング[A]【講師A】", "1", "201"
        sheet["B9"], sheet["C9"], sheet["D9"] = "Webプログラミング[A]【講師A】", "1", "201"
        sheet["A9"].border = bottom

        sheet["A10"] = "3限 13:30-15:00"
        sheet["A10"].border = bottom
        sheet["A11"] = "4限 15:10-16:40"
        sheet["B11"], sheet["C11"], sheet["D11"] = "企業説明会", "全学年", "ホール"
        sheet["H11"], sheet["I11"], sheet["J11"] = "企業説明会", "1", "下記参照"
        sheet["A11"].border = bottom
        sheet["B12"] = "運営上の連絡事項"
        sheet["H12"], sheet["I12"], sheet["J12"] = "A組：別会場", "", "301"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def analyze_classified(self):
        return self.service.analyze(
            self.classified_workbook_bytes(),
            "匿名化時間割.xlsx",
            "2025-04-21",
            "2025-04-27",
            sheet_name="2025通年時間割",
        )

    @staticmethod
    def subject_workbook_bytes(subjects: list[str]) -> io.BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2025通年時間割"
        sheet["B1"] = date(2025, 4, 21)
        for offset, column in enumerate(("E", "H", "K", "N", "Q", "T"), start=1):
            sheet[f"{column}1"] = f"=B1+{offset}"
        sheet["A2"] = "1限 09:20-10:50"
        bottom = Border(bottom=Side(style="thin"))
        for offset, subject in enumerate(subjects):
            row = 2 + offset
            sheet.cell(row=row, column=2, value=subject)
            sheet.cell(row=row, column=3, value="2")
            sheet.cell(row=row, column=4, value="701")
        sheet.cell(row=1 + len(subjects), column=1).border = bottom
        for period, row, timing in (
            (2, 3 + len(subjects), "11:00-12:30"),
            (3, 4 + len(subjects), "13:30-15:00"),
            (4, 5 + len(subjects), "15:10-16:40"),
        ):
            sheet.cell(row=row, column=1, value=f"{period}限 {timing}").border = bottom
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def analyze_subjects(self, subjects: list[str]):
        return self.service.analyze(
            self.subject_workbook_bytes(subjects),
            "科目フロー.xlsx",
            "2025-04-21",
            "2025-04-27",
            sheet_name="2025通年時間割",
        )

    def test_subject_grade_and_room_normalization_preserves_source_meaning(self):
        parsed = self.service.parse_subject("ドローン概論（A組）[A][C]【CDP・専任】")
        self.assertEqual(parsed["subjectName"], "ドローン概論[A][C]")
        self.assertEqual(parsed["subjectBaseName"], "ドローン概論")
        self.assertEqual(parsed["courseVariantTags"], ["A", "C"])
        self.assertEqual(parsed["groupTags"], ["A組"])
        self.assertEqual(parsed["instructors"], ["CDP", "専任"])
        self.assertEqual(self.service.parse_grades("全学年"), ([1, 2, 3], "全学年"))
        self.assertEqual(self.service.parse_grades("2,3"), ([2, 3], "2,3"))
        self.assertIsNone(self.service.parse_grades("研究科")[0])
        self.assertEqual(self.service.parse_rooms(721722, "000,000"), (["721", "722"], None))
        rooms, warning = self.service.parse_rooms(721.722)
        self.assertEqual(rooms, ["721.722"])
        self.assertIn("小数点", warning)

    def test_formula_dates_merged_cells_periods_and_parallel_lessons_are_parsed(self):
        preview = self.analyze()
        self.assertEqual(preview["blockingCount"], 0)
        self.assertEqual(preview["entryCount"], 5)
        self.assertEqual(preview["dateCount"], 7)
        token_preview = self.service.previews[preview["token"]]
        first, second = token_preview.entries[:2]
        self.assertEqual((first["date"], first["period"]), ("2025-04-21", 1))
        self.assertEqual(first["grades"], [1])
        self.assertEqual(first["rooms"], ["721", "722"])
        self.assertEqual(second["grades"], [1])
        self.assertEqual(second["rooms"], ["721", "722"])
        web = next(item for item in token_preview.entries if item["subjectName"] == "Webプログラミング")
        self.assertEqual(web["subjectRaw"], "Webプログラミング（A組）【CDP】")
        exam = next(item for item in token_preview.entries if item["kind"] == "exam")
        self.assertEqual(exam["status"], "tentative")
        event = next(item for item in token_preview.entries if item["kind"] == "event")
        self.assertIsNone(event["courseId"])
        self.assertEqual(token_preview.periods, DEFAULT_PERIODS)

    def test_unknown_subject_blocks_and_suspicious_room_warns_without_repair(self):
        preview = self.analyze(unknown_subject=True, suspicious_room=True)
        codes = {item["code"] for item in preview["issues"]}
        self.assertIn("unregistered_subject", codes)
        self.assertIn("suspicious_room", codes)
        self.assertGreater(preview["blockingCount"], 0)
        entry = next(item for item in self.service.previews[preview["token"]].entries if item["rooms"] == ["721.722"])
        self.assertEqual(entry["rooms"], ["721.722"])
        with self.assertRaises(TimetableError) as context:
            self.service.save_preview(preview["token"], True)
        self.assertEqual(context.exception.code, "blocking_issues")

    def test_diff_detects_changes_and_ignores_row_order(self):
        base = {
            "date": "2025-09-04", "period": 1, "courseId": "java", "subjectName": "Java",
            "subjectRaw": "Java", "grades": [1], "gradeRaw": "1", "groupTags": ["A"],
            "instructors": ["山田"], "rooms": ["721"], "kind": "class", "status": "confirmed", "entryId": "tt-0000000000000000",
        }
        other = {**base, "subjectName": "Webプログラミング", "subjectRaw": "Webプログラミング", "courseId": "web-programming", "groupTags": ["C"], "entryId": "tt-1111111111111111"}
        changed = {**base, "rooms": ["733", "734"], "instructors": ["佐藤"], "kind": "exam", "status": "tentative"}
        changes = self.service.compare([base, other], [other, changed])
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["type"], "changed")
        self.assertEqual(changes[0]["fields"], ["教室", "講師", "授業種別", "確定状態"])

    def test_confirmed_range_preserves_outside_snapshot_and_warns_on_blanking(self):
        previous = {
            "date": "2025-04-21", "period": 4, "courseId": "java", "subjectRaw": "Java",
            "subjectName": "Java", "grades": [1], "gradeRaw": "1", "groupTags": [],
            "instructors": [], "rooms": [], "kind": "class", "status": "confirmed", "entryId": "tt-aaaaaaaaaaaaaaaa",
        }
        outside = {**previous, "date": "2025-05-01", "entryId": "tt-bbbbbbbbbbbbbbbb"}
        self.service.snapshot_path.write_text(
            json.dumps({"entries": [previous, outside]}, ensure_ascii=False), encoding="utf-8"
        )
        preview = self.analyze()
        parsed = self.service.previews[preview["token"]]
        self.assertTrue(any(item["date"] == "2025-05-01" for item in parsed.merged_snapshot))
        self.assertTrue(any(item["code"] == "previous_entry_now_blank" for item in preview["issues"]))

    def test_save_splits_months_validates_schema_and_creates_backup(self):
        preview = self.analyze()
        result = self.service.save_preview(preview["token"], True)
        self.assertTrue(result["success"])
        manifest = json.loads((self.repo / "data" / "timetable" / "manifest.json").read_text(encoding="utf-8"))
        month = json.loads((self.repo / "data" / "timetable" / "2025-04.json").read_text(encoding="utf-8"))
        periods = json.loads((self.repo / "data" / "timetable" / "periods.json").read_text(encoding="utf-8"))
        self.assertEqual(manifest["months"], ["2025-04"])
        self.assertEqual(len(month["days"]["2025-04-21"]), 5)
        self.assertNotIn("date", month["days"]["2025-04-21"][0])
        self.assertEqual(periods["periods"], DEFAULT_PERIODS)
        for document, schema_name in (
            (manifest, "timetable-manifest.schema.json"),
            (month, "timetable-month.schema.json"),
            (periods, "timetable-periods.schema.json"),
        ):
            schema = json.loads((self.repo / "data" / schema_name).read_text(encoding="utf-8"))
            Draft202012Validator(schema).validate(document)
        self.assertTrue(Path(result["backupPath"]).is_dir())

    def test_private_write_failure_restores_public_and_snapshot(self):
        self.service.public_root.mkdir(parents=True)
        old_manifest = {"old": True}
        (self.service.public_root / "manifest.json").write_text(
            json.dumps(old_manifest), encoding="utf-8"
        )
        old_snapshot = {"entries": [], "marker": "old"}
        self.service.snapshot_path.write_text(json.dumps(old_snapshot), encoding="utf-8")
        preview = self.analyze()
        original_atomic = self.service._atomic_json

        def fail_metadata(path, document):
            if path == self.service.metadata_path:
                raise OSError("simulated write failure")
            return original_atomic(path, document)

        with patch.object(self.service, "_atomic_json", side_effect=fail_metadata):
            with self.assertRaises(TimetableError) as context:
                self.service.save_preview(preview["token"], True)

        self.assertEqual(context.exception.code, "write_failed")
        restored_manifest = json.loads((self.service.public_root / "manifest.json").read_text(encoding="utf-8"))
        restored_snapshot = json.loads(self.service.snapshot_path.read_text(encoding="utf-8"))
        self.assertEqual(restored_manifest, old_manifest)
        self.assertEqual(restored_snapshot, old_snapshot)

    def test_period_block_inheritance_recovers_rows_without_repeated_period_numbers(self):
        preview = self.analyze_classified()
        parsed = self.service.previews[preview["token"]]
        inherited = next(item for item in parsed.entries if item["date"] == "2025-04-22")
        self.assertEqual(inherited["period"], 1)
        self.assertEqual(inherited["subjectName"], "Java[Z]")
        self.assertNotIn("period_parse_failed", {item["code"] for item in preview["issues"]})

    def test_structural_headers_are_ignored_and_ordinary_notes_are_references(self):
        preview = self.analyze_classified()
        self.assertFalse(any(item.get("subjectRaw") == "科目【講師】" for item in preview["issues"]))
        note = next(item for item in preview["issues"] if item.get("subjectRaw") == "運営上の連絡事項")
        self.assertEqual((note["severity"], note["code"]), ("reference", "supplemental_info"))

    def test_see_below_keeps_linked_supplement_visible_without_publishing_it(self):
        preview = self.analyze_classified()
        self.assertTrue(any(item["code"] == "see_below" and item["severity"] == "warning" for item in preview["issues"]))
        detail = next(item for item in preview["issues"] if item.get("subjectRaw") == "A組：別会場")
        self.assertEqual(detail.get("referenceCategory"), "see_below_detail")
        entries = self.service.previews[preview["token"]].entries
        self.assertFalse(any(item["subjectRaw"] == "A組：別会場" for item in entries))

    def test_clear_special_event_is_an_entry_without_course_mapping(self):
        preview = self.analyze_classified()
        entries = self.service.previews[preview["token"]].entries
        event = next(item for item in entries if item["subjectName"] == "企業説明会")
        self.assertEqual(event["kind"], "event")
        self.assertIsNone(event["courseId"])
        self.assertFalse(any(item["code"] == "unregistered_subject" and item.get("subjectRaw") == "企業説明会" for item in preview["issues"]))

    def test_duplicate_detection_keeps_parallel_groups_and_grades(self):
        base = {
            "date": "2025-04-21", "period": 1, "courseId": "java", "subjectName": "Java",
            "grades": [1], "groupTags": ["A"], "instructors": ["講師"], "rooms": ["101"],
            "kind": "class", "status": "confirmed", "subjectRaw": "Java[A]",
        }
        issues = []
        self.service._detect_duplicates([
            base,
            {**base, "groupTags": ["C"]},
            {**base, "grades": [2]},
            {**base, "subjectName": "Webプログラミング", "courseId": "web-programming"},
        ], issues)
        self.assertEqual(issues, [])

    def test_exact_duplicate_remains_a_warning(self):
        entry = {
            "date": "2025-04-21", "period": 1, "courseId": "java", "subjectName": "Java",
            "grades": [1], "groupTags": ["A"], "instructors": ["講師"], "rooms": ["101"],
            "kind": "class", "status": "confirmed", "subjectRaw": "Java[A]",
        }
        issues = []
        self.service._detect_duplicates([entry, dict(entry)], issues)
        self.assertEqual(len(issues), 1)
        self.assertEqual((issues[0]["severity"], issues[0]["code"]), ("warning", "duplicate_entry"))

    def test_preview_separates_blocking_warning_and_reference_counts(self):
        preview = self.analyze_classified()
        self.assertGreater(preview["blockingCount"], 0)
        self.assertGreater(preview["warningCount"], 0)
        self.assertEqual(preview["referenceCount"], 2)
        self.assertTrue(any(item["severity"] == "reference" for item in preview["issueCounts"]))

    def test_symbol_types_preserve_brackets_and_meanings(self):
        variant = self.service.parse_subject("WEB制作[A]【講師X】")
        self.assertEqual(variant["subjectName"], "WEB制作[A]")
        self.assertEqual(variant["subjectBaseName"], "WEB制作")
        self.assertEqual(variant["courseVariantTags"], ["A"])
        self.assertEqual(variant["groupTags"], [])
        self.assertEqual(variant["instructors"], ["講師X"])

        ascii_group = self.service.parse_subject("WEB制作(A)【講師X】")
        full_group = self.service.parse_subject("WEB制作（A組）【講師X】")
        composite = self.service.parse_subject("WEB制作[A]（A組）【講師X】")
        self.assertEqual(ascii_group["groupTags"], ["A"])
        self.assertEqual(ascii_group["courseVariantTags"], [])
        self.assertEqual(full_group["groupTags"], ["A組"])
        self.assertEqual(composite["subjectName"], "WEB制作[A]")
        self.assertEqual(composite["courseVariantTags"], ["A"])
        self.assertEqual(composite["groupTags"], ["A組"])
        self.assertEqual(composite["instructors"], ["講師X"])

        token_types = {
            item["rawToken"]: item["tokenType"]
            for item in self.service.parse_subject("WEB制作[A](A)【A】")["tokens"]
        }
        self.assertEqual(token_types, {"[A]": "course-variant", "(A)": "group", "【A】": "instructor"})

    def test_ordinary_parentheses_remain_part_of_subject(self):
        parsed = self.service.parse_subject("WEB制作（基礎）")
        self.assertEqual(parsed["subjectName"], "WEB制作(基礎)")
        self.assertEqual(parsed["subjectBaseName"], "WEB制作(基礎)")
        self.assertEqual(parsed["groupTags"], [])
        qualifier = next(item for item in parsed["tokens"] if item["rawToken"] == "（基礎）")
        self.assertEqual(qualifier["tokenType"], "subject-qualifier")

    def test_round_group_aliases_can_share_a_canonical_group(self):
        first = self.service.save_group_mapping("(A)", group_id="group-a")
        second = self.service.save_group_mapping("（A組）", group_id="group-a")
        config = self.service._config()
        self.assertEqual(first["alias"], "(A)")
        self.assertEqual(second["alias"], "(A組)")
        self.assertEqual(config["groupMappings"]["(A)"], "group-a")
        self.assertEqual(config["groupMappings"]["(A組)"], "group-a")
        self.assertEqual(config["legacyGroupTags"], {"A": "選択A", "C": "選択C"})

    def test_legacy_bare_group_mapping_is_not_reused(self):
        config = self.service._config()
        self.assertNotIn("(A)", config["groupMappings"])
        self.assertEqual(config["legacyGroupTags"]["A"], "選択A")

    def test_course_mapping_keys_keep_variant_information(self):
        self.service.save_mapping("WEB制作[A]", "web-programming")
        self.service.save_mapping("WEB制作[B]", "java")
        mappings = self.service._config()["subjectMappings"]
        self.assertEqual(mappings["WEB制作[A]"], "web-programming")
        self.assertEqual(mappings["WEB制作[B]"], "java")

    def test_unregistered_variants_are_separate_but_grouped_by_base_name(self):
        preview = self.analyze_subjects(["After Effects[A]【講師X】", "After Effects[C]【講師X】"])
        subjects = preview["unregisteredSubjects"]
        self.assertEqual([item["subjectName"] for item in subjects], ["After Effects[A]", "After Effects[C]"])
        self.assertEqual({item["subjectBaseName"] for item in subjects}, {"After Effects"})
        self.assertEqual([item["courseVariantTags"] for item in subjects], [["A"], ["C"]])
        self.assertEqual([item["count"] for item in subjects], [1, 1])

    def test_exact_candidate_never_uses_base_name_as_an_automatic_match(self):
        document = json.loads(self.service.courses_path.read_text(encoding="utf-8"))
        document["courses"].extend([
            {"id": "after-effects", "title": "After Effects"},
            {"id": "generative-ai", "title": "生成AI"},
        ])
        self.service.courses_path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")
        preview = self.analyze_subjects(["After Effects[A]【講師X】", "生成AI【講師Y】"])
        by_name = {item["subjectName"]: item for item in preview["unregisteredSubjects"]}
        self.assertEqual(by_name["After Effects[A]"]["exactCourseCandidates"], [])
        self.assertEqual(by_name["After Effects[A]"]["relatedBaseCourses"][0]["id"], "after-effects")
        self.assertEqual(by_name["生成AI"]["exactCourseCandidates"][0]["id"], "generative-ai")
        self.assertTrue(all(entry["courseId"] is None for entry in self.service.previews[preview["token"]].entries))

    def test_special_mapping_allows_a_course_without_course_id_and_is_reused(self):
        preview = self.analyze_subjects(["スーツ講座【外部講師】"])
        self.assertEqual(preview["blockingCount"], 1)

        self.service.save_item_mapping("スーツ講座", "special")
        refreshed = self.service.refresh_preview(preview["token"])
        entry = self.service.previews[preview["token"]].entries[0]
        self.assertEqual(entry["kind"], "special")
        self.assertIsNone(entry["courseId"])
        self.assertEqual(refreshed["blockingCount"], 0)
        self.assertEqual(refreshed["unregisteredSubjects"], [])

        documents = self.service._public_documents(
            self.service.previews[preview["token"]].merged_snapshot,
            self.service.previews[preview["token"]].periods,
        )
        month = documents["2025-04.json"]
        public_entry = month["days"]["2025-04-21"][0]
        self.assertEqual(public_entry["kind"], "special")
        self.assertIsNone(public_entry["courseId"])
        self.assertNotIn("_sourceCell", public_entry)
        self.service._validate_public_document("2025-04.json", month)

        reanalyzed = self.analyze_subjects(["スーツ講座【外部講師】"])
        self.assertFalse(any(
            issue["code"] == "unregistered_subject" for issue in reanalyzed["issues"]
        ))

    def test_event_holiday_and_other_explicit_mappings_do_not_require_courses(self):
        preview = self.analyze_subjects(["キャリア交流", "学年間休み", "校内予定"])
        self.service.save_item_mapping("キャリア交流", "event")
        self.service.save_item_mapping("学年間休み", "holiday")
        self.service.save_item_mapping("校内予定", "other")
        refreshed = self.service.refresh_preview(preview["token"])
        entries = {
            entry["subjectName"]: entry
            for entry in self.service.previews[preview["token"]].entries
        }
        self.assertEqual(entries["キャリア交流"]["kind"], "event")
        self.assertEqual(entries["学年間休み"]["kind"], "holiday")
        self.assertEqual(entries["校内予定"]["kind"], "other")
        self.assertTrue(all(entry["courseId"] is None for entry in entries.values()))
        self.assertEqual(refreshed["blockingCount"], 0)

    def test_unclassified_class_stays_blocking_and_schema_requires_course_id(self):
        preview = self.analyze_subjects(["After Effects[A]【講師X】"])
        entry = self.service.previews[preview["token"]].entries[0]
        self.assertEqual(entry["kind"], "class")
        self.assertIsNone(entry["courseId"])
        self.assertEqual(preview["blockingCount"], 1)

        invalid_public = self.service._public_documents(
            self.service.previews[preview["token"]].merged_snapshot,
            self.service.previews[preview["token"]].periods,
        )["2025-04.json"]
        with self.assertRaises(TimetableError) as raised:
            self.service._validate_public_document("2025-04.json", invalid_public)
        self.assertEqual(raised.exception.code, "schema_error")

    def test_course_and_item_mappings_cannot_conflict_without_confirmed_change(self):
        self.service.save_item_mapping("スーツ講座", "special")
        with self.assertRaises(TimetableError) as raised:
            self.service.save_mapping("スーツ講座", "java")
        self.assertEqual(raised.exception.code, "classification_conflict")

        self.service.save_mapping("スーツ講座", "java", replace_existing=True)
        config = self.service._config()
        self.assertEqual(config["subjectMappings"]["スーツ講座"], "java")
        self.assertNotIn("スーツ講座", config["timetableItemMappings"])

        with self.assertRaises(TimetableError) as reverse:
            self.service.save_item_mapping("スーツ講座", "event")
        self.assertEqual(reverse.exception.code, "classification_conflict")
        self.service.save_item_mapping("スーツ講座", "event", replace_existing=True)
        config = self.service._config()
        self.assertNotIn("スーツ講座", config["subjectMappings"])
        self.assertEqual(config["timetableItemMappings"]["スーツ講座"]["type"], "event")

    def test_changing_special_to_class_requires_a_course(self):
        preview = self.analyze_subjects(["スーツ講座"])
        self.service.save_item_mapping("スーツ講座", "special")
        refreshed = self.service.refresh_preview(preview["token"])
        with self.assertRaises(TimetableError) as raised:
            self.service.save_item_mapping("スーツ講座", "class", replace_existing=True)
        self.assertEqual(raised.exception.code, "course_required")
        special_setting = next(
            item for item in refreshed["mappingSettings"] if item["subjectName"] == "スーツ講座"
        )
        self.assertEqual(special_setting["classificationLabel"], "単発講座・特別授業")

        self.service.save_mapping("スーツ講座", "java", replace_existing=True)
        changed = self.service.refresh_preview(preview["token"])
        entry = self.service.previews[preview["token"]].entries[0]
        self.assertEqual(entry["kind"], "class")
        self.assertEqual(entry["courseId"], "java")
        class_setting = next(
            item for item in changed["mappingSettings"] if item["subjectName"] == "スーツ講座"
        )
        self.assertEqual(class_setting["classificationLabel"], "通常授業")

    def test_item_mapping_route_refreshes_preview_and_reports_write_failure(self):
        application = create_app(self.repo, self.work)
        timetable = application.config["TIMETABLE_SERVICE"]
        preview = timetable.analyze(
            self.subject_workbook_bytes(["スーツ講座"]),
            "分類テスト.xlsx",
            "2025-04-21",
            "2025-04-27",
            sheet_name="2025通年時間割",
        )
        response = application.test_client().post("/api/timetable/mappings/item", json={
            "subjectName": "スーツ講座",
            "classification": "special",
            "previewToken": preview["token"],
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["preview"]["blockingCount"], 0)

        with patch.object(
            timetable,
            "save_item_mapping",
            side_effect=TimetableError("classification write failed", code="write_failed", status=500),
        ):
            failed = application.test_client().post("/api/timetable/mappings/item", json={
                "subjectName": "別の講座", "classification": "special",
            })
        self.assertEqual(failed.status_code, 500)
        self.assertEqual(failed.get_json()["error"], "classification write failed")

    def test_course_context_maps_only_the_created_variant_and_restores_preview(self):
        preview = self.analyze_subjects(["After Effects[A]【講師X】", "After Effects[C]【講師X】"])
        context = self.service.create_course_context(preview["token"], "After Effects[A]")
        document = json.loads(self.service.courses_path.read_text(encoding="utf-8"))
        document["courses"].append({"id": "after-effects-a", "title": "After Effects[A]"})
        self.service.courses_path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")

        result = self.service.complete_course_registration(context["token"], "after-effects-a")
        restored = self.service.get_preview(preview["token"])
        mappings = self.service._config()["subjectMappings"]
        self.assertEqual(mappings["AfterEffects[A]"], "after-effects-a")
        self.assertNotIn("AfterEffects[C]", mappings)
        self.assertEqual(
            [item["subjectName"] for item in restored["unregisteredSubjects"]],
            ["After Effects[C]"],
        )
        self.assertEqual(result["returnUrl"], f"/timetable?preview={preview['token']}")
        self.assertEqual(restored["blockingCount"], preview["blockingCount"] - 1)

    def test_cancelled_course_context_creates_no_mapping(self):
        preview = self.analyze_subjects(["After Effects[A]【講師X】"])
        context = self.service.create_course_context(preview["token"], "After Effects[A]")
        result = self.service.cancel_course_context(context["token"])
        self.assertEqual(result["returnUrl"], f"/timetable?preview={preview['token']}")
        self.assertNotIn("AfterEffects[A]", self.service._config()["subjectMappings"])

    def test_mapping_to_deleted_or_archived_course_is_blocking(self):
        config = self.service._config()
        config["subjectMappings"]["AfterEffects[A]"] = "missing-course"
        self.service.config_path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
        preview = self.analyze_subjects(["After Effects[A]【講師X】"])
        issue = next(item for item in preview["issues"] if item["code"] == "mapped_course_missing")
        self.assertEqual(issue["courseId"], "missing-course")
        self.assertEqual(preview["blockingCount"], 1)

    def test_register_route_maps_only_after_course_save_succeeds(self):
        application = create_app(self.repo, self.work)
        importer = application.config["IMPORTER_SERVICE"]
        timetable = application.config["TIMETABLE_SERVICE"]
        publisher = application.config["PUBLISHER_SERVICE"]
        saved = {
            "id": "after-effects-a", "title": "After Effects[A]",
            "coursesPath": "data/courses.json", "backupPath": "backup.json",
        }
        with (
            patch.object(importer, "register_course", return_value=saved),
            patch.object(timetable, "complete_course_registration", return_value={
                "returnUrl": "/timetable?preview=preview-1",
            }) as complete,
            patch.object(publisher, "record_operation"),
        ):
            response = application.test_client().post("/api/register", json={
                "preparationToken": "preparation",
                "validationToken": "validation",
                "course": {"id": "after-effects-a"},
                "timetableContextToken": "context-a",
            })
        self.assertEqual(response.status_code, 200)
        complete.assert_called_once_with("context-a", "after-effects-a")

        with (
            patch.object(importer, "register_course", side_effect=ImporterError("course save failed")),
            patch.object(timetable, "complete_course_registration") as complete,
            patch.object(publisher, "record_operation"),
        ):
            failed = application.test_client().post("/api/register", json={
                "preparationToken": "preparation",
                "validationToken": "validation",
                "course": {"id": "after-effects-c"},
                "timetableContextToken": "context-c",
            })
        self.assertEqual(failed.status_code, 400)
        complete.assert_not_called()

    def test_register_route_reports_mapping_failure_after_course_was_created(self):
        application = create_app(self.repo, self.work)
        importer = application.config["IMPORTER_SERVICE"]
        timetable = application.config["TIMETABLE_SERVICE"]
        publisher = application.config["PUBLISHER_SERVICE"]
        saved = {
            "id": "after-effects-a", "title": "After Effects[A]",
            "coursesPath": "data/courses.json", "backupPath": "backup.json",
        }
        with (
            patch.object(importer, "register_course", return_value=saved),
            patch.object(
                timetable, "complete_course_registration",
                side_effect=TimetableError("mapping write failed", code="write_failed", status=500),
            ),
            patch.object(publisher, "record_operation"),
        ):
            response = application.test_client().post("/api/register", json={
                "preparationToken": "preparation",
                "validationToken": "validation",
                "course": {"id": "after-effects-a"},
                "timetableContextToken": "context-a",
            })
        payload = response.get_json()
        self.assertEqual(response.status_code, 409)
        self.assertTrue(payload["courseCreated"])
        self.assertEqual(payload["code"], "timetable_mapping_failed")
        self.assertIn("授業", payload["error"])
        self.assertIn("時間割との対応", payload["error"])


if __name__ == "__main__":
    unittest.main()
