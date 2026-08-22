"""Safe Excel-to-JSON timetable import for the local ClassView admin tool."""

from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import shutil
import tempfile
import threading
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, BinaryIO
from uuid import uuid4

from jsonschema import Draft202012Validator
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.datetime import from_excel


DAY_COLUMNS = (2, 5, 8, 11, 14, 17, 20)
DEFAULT_PERIODS = {
    "1": {"start": "09:20", "end": "10:50"},
    "2": {"start": "11:00", "end": "12:30"},
    "3": {"start": "13:30", "end": "15:00"},
    "4": {"start": "15:10", "end": "16:40"},
}
KINDS = {"class", "exam", "special", "event", "holiday", "other"}
TIMETABLE_ITEM_KINDS = {"exam", "special", "event", "holiday", "other"}
KIND_LABELS = {
    "class": "通常授業",
    "exam": "試験",
    "special": "単発講座・特別授業",
    "event": "行事・説明会",
    "holiday": "休暇",
    "other": "その他",
}
STATUSES = {"confirmed", "tentative"}
ISSUE_LABELS = {
    "date_parse_failed": "日付を解析できない箇所",
    "duplicate_entry": "重複の可能性",
    "empty_exam_name": "科目名のない試験",
    "inconsistent_period_time": "時限時刻の不一致",
    "invalid_grade": "学年を解析できない箇所",
    "mapped_course_missing": "対応先授業が見つからない科目",
    "no_entries_in_range": "指定期間内のデータ不足",
    "period_parse_failed": "時限を特定できない予定",
    "period_times_missing": "時限時刻の不足",
    "previous_entry_now_blank": "前回から空白になった予定",
    "see_below": "「下記参照」の確認",
    "supplemental_info": "Excel内の補足記載",
    "suspicious_room": "確認が必要な教室表記",
    "tentative": "候補日・予定",
    "unexpected_structure": "Excel構造の不一致",
    "unknown_group_tag": "未登録グループ",
    "unregistered_subject": "未登録科目",
}


class TimetableError(Exception):
    def __init__(self, message: str, *, code: str = "timetable_error", status: int = 400):
        super().__init__(message)
        self.message = message
        self.code = code
        self.status = status


@dataclass
class TimetablePreview:
    token: str
    created_at: datetime
    filename: str
    sheet_name: str
    start_date: date
    end_date: date
    source_modified_at: str | None
    entries: list[dict[str, Any]]
    periods: dict[str, dict[str, str]]
    issues: list[dict[str, Any]]
    changes: list[dict[str, Any]]
    merged_snapshot: list[dict[str, Any]]
    token_summary: dict[str, list[dict[str, Any]]]
    token_occurrences: list[dict[str, Any]]
    comparison_base: list[dict[str, Any]]
    date_count: int


@dataclass
class PendingCourseContext:
    token: str
    preview_token: str
    subject_name: str
    created_at: datetime


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return unicodedata.normalize("NFKC", str(value)).strip()


def _raw_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _mapping_key(value: str) -> str:
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", value)).strip()


def _iso_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


class TimetableService:
    """Parse a workbook, preview changes, then commit validated public JSON."""

    def __init__(self, repo_root: Path, work_root: Path):
        self.repo_root = Path(repo_root).resolve()
        self.work_root = Path(work_root).resolve()
        self.data_root = self.repo_root / "data"
        self.public_root = self.data_root / "timetable"
        self.runtime_root = self.work_root / "runtime" / "timetable"
        self.backups_root = self.work_root / "backups" / "timetable"
        self.config_path = self.runtime_root / "config.json"
        self.snapshot_path = self.runtime_root / "last-import.json"
        self.metadata_path = self.runtime_root / "import-metadata.json"
        self.courses_path = self.data_root / "courses.json"
        self.month_schema_path = self.data_root / "timetable-month.schema.json"
        self.manifest_schema_path = self.data_root / "timetable-manifest.schema.json"
        self.periods_schema_path = self.data_root / "timetable-periods.schema.json"
        self.runtime_root.mkdir(parents=True, exist_ok=True)
        self.backups_root.mkdir(parents=True, exist_ok=True)
        self.previews: dict[str, TimetablePreview] = {}
        self.course_contexts: dict[str, PendingCourseContext] = {}
        self.lock = threading.RLock()

    def _load_json(self, path: Path, default: Any) -> Any:
        if not path.exists():
            return copy.deepcopy(default)
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise TimetableError(
                f"{path.name} を読み込めません。管理担当者へ確認してください。",
                code="invalid_timetable_state",
                status=500,
            ) from error

    def _config(self) -> dict[str, Any]:
        value = self._load_json(
            self.config_path,
            {
                "sheetName": "2025通年時間割",
                "subjectMappings": {},
                "timetableItemMappings": {},
                "canonicalGroups": {},
                "groupMappings": {},
                "legacyGroupTags": {},
            },
        )
        if not isinstance(value, dict):
            raise TimetableError("時間割の内部設定が不正です。", status=500)
        value.setdefault("sheetName", "2025通年時間割")
        value.setdefault("subjectMappings", {})
        value.setdefault("timetableItemMappings", {})
        value.setdefault("canonicalGroups", {})
        value.setdefault("groupMappings", {})
        legacy_group_tags = value.setdefault("legacyGroupTags", {})
        old_group_tags = value.get("groupTags", {})
        if isinstance(old_group_tags, dict):
            # Old keys such as "A" no longer contain enough information to
            # distinguish [A] from (A). Preserve them for audit, but never use
            # them as aliases in the new parser.
            for key, display_name in old_group_tags.items():
                legacy_group_tags.setdefault(key, display_name)
        if not all(isinstance(value[key], dict) for key in (
            "subjectMappings", "timetableItemMappings", "canonicalGroups",
            "groupMappings", "legacyGroupTags"
        )):
            raise TimetableError("時間割の対応設定が不正です。", status=500)
        for key, mapping in value["timetableItemMappings"].items():
            if (
                not isinstance(key, str)
                or not isinstance(mapping, dict)
                or mapping.get("type") not in TIMETABLE_ITEM_KINDS
                or not _text(mapping.get("displayName"))
            ):
                raise TimetableError("時間割項目の分類設定が不正です。", status=500)
        conflicts = set(value["subjectMappings"]) & set(value["timetableItemMappings"])
        if conflicts:
            raise TimetableError(
                "通常授業との対応と時間割項目の分類が重複しています。管理担当者へ確認してください。",
                code="classification_conflict",
                status=500,
            )
        return value

    def status(self) -> dict[str, Any]:
        config = self._config()
        metadata = self._load_json(self.metadata_path, {})
        snapshot = self._load_json(self.snapshot_path, {"entries": []})
        entries = snapshot.get("entries", []) if isinstance(snapshot, dict) else []
        return {
            "sheetName": config["sheetName"],
            "lastFilename": metadata.get("sourceFilename"),
            "lastImportedAt": metadata.get("lastImportedAt"),
            "sourceModifiedAt": metadata.get("sourceModifiedAt"),
            "entryCount": len(entries) if isinstance(entries, list) else 0,
            "subjectMappings": config["subjectMappings"],
            "timetableItemMappings": config["timetableItemMappings"],
            "canonicalGroups": self._canonical_group_options(config),
            "groupMappings": config["groupMappings"],
            "legacyGroupTags": config["legacyGroupTags"],
            "courses": self._course_options(),
        }

    def _course_options(self) -> list[dict[str, str | None]]:
        document = self._load_json(self.courses_path, {"courses": []})
        options = []
        for course in document.get("courses", []):
            if isinstance(course, dict) and isinstance(course.get("id"), str):
                options.append({
                    "id": course["id"],
                    "title": _text(course.get("title")),
                    "academicYear": _text(course.get("academicYear")) or None,
                })
        return sorted(options, key=lambda item: (item["title"], item["id"]))

    def save_mapping(
        self, subject_name: str, course_id: str, *, replace_existing: bool = False
    ) -> dict[str, Any]:
        key = _mapping_key(subject_name)
        valid_ids = {item["id"] for item in self._course_options()}
        if not key:
            raise TimetableError("Excel上の科目名を入力してください。")
        if course_id not in valid_ids:
            raise TimetableError("選択した授業IDは現在の授業一覧にありません。")
        with self.lock:
            config = self._config()
            existing_item = config["timetableItemMappings"].get(key)
            if existing_item and not replace_existing:
                raise TimetableError(
                    f"「{subject_name}」は現在「{KIND_LABELS[existing_item['type']]}」です。分類変更を確認してください。",
                    code="classification_conflict",
                    status=409,
                )
            if existing_item:
                config["timetableItemMappings"].pop(key, None)
            config["subjectMappings"][key] = course_id
            self._backup_private_config()
            self._atomic_json(self.config_path, config)
        return {"success": True, "subjectName": subject_name, "courseId": course_id}

    def save_item_mapping(
        self, subject_name: str, item_kind: str, *, replace_existing: bool = False
    ) -> dict[str, Any]:
        display_name = _text(subject_name)
        key = _mapping_key(display_name)
        normalized_kind = _text(item_kind).lower()
        if not key:
            raise TimetableError("Excel上の項目名を入力してください。")
        if normalized_kind == "class":
            raise TimetableError(
                "通常授業として扱う場合は、既存授業との対応または新規授業登録が必要です。",
                code="course_required",
                status=409,
            )
        if normalized_kind not in TIMETABLE_ITEM_KINDS:
            raise TimetableError("時間割項目の種類を選択してください。")
        with self.lock:
            config = self._config()
            existing_course = config["subjectMappings"].get(key)
            existing_item = config["timetableItemMappings"].get(key)
            changing_item = existing_item and existing_item.get("type") != normalized_kind
            if (existing_course or changing_item) and not replace_existing:
                current = "通常授業" if existing_course else KIND_LABELS[existing_item["type"]]
                raise TimetableError(
                    f"「{display_name}」は現在「{current}」です。分類変更を確認してください。",
                    code="classification_conflict",
                    status=409,
                )
            if existing_course:
                config["subjectMappings"].pop(key, None)
            config["timetableItemMappings"][key] = {
                "type": normalized_kind,
                "displayName": display_name,
                "updatedAt": _iso_now(),
            }
            self._backup_private_config()
            self._atomic_json(self.config_path, config)
        return {
            "success": True,
            "subjectName": display_name,
            "classification": normalized_kind,
            "classificationLabel": KIND_LABELS[normalized_kind],
        }

    @staticmethod
    def _group_alias(raw_token: str) -> str:
        normalized = re.sub(r"\s+", "", unicodedata.normalize("NFKC", raw_token)).strip()
        if not re.fullmatch(r"\([^()]+\)", normalized):
            raise TimetableError("グループの元表記は丸括弧付きで指定してください。")
        return normalized

    @staticmethod
    def _canonical_group_options(config: dict[str, Any]) -> list[dict[str, str]]:
        return sorted(
            [
                {"id": group_id, "displayName": _text(value.get("displayName"))}
                for group_id, value in config["canonicalGroups"].items()
                if isinstance(value, dict) and _text(value.get("displayName"))
            ],
            key=lambda item: (item["displayName"], item["id"]),
        )

    def save_group_mapping(
        self, raw_token: str, *, group_id: str = "", display_name: str = ""
    ) -> dict[str, Any]:
        alias = self._group_alias(raw_token)
        normalized_group_id = _mapping_key(group_id)
        normalized_display_name = _text(display_name)
        if bool(normalized_group_id) == bool(normalized_display_name):
            raise TimetableError("既存グループを選ぶか、新しいグループ名を入力してください。")
        with self.lock:
            config = self._config()
            if normalized_group_id:
                if normalized_group_id not in config["canonicalGroups"]:
                    raise TimetableError("選択したグループは現在の設定にありません。")
                canonical_id = normalized_group_id
            else:
                digest = hashlib.sha256(normalized_display_name.encode("utf-8")).hexdigest()[:12]
                canonical_id = f"group-{digest}"
                config["canonicalGroups"].setdefault(
                    canonical_id, {"displayName": normalized_display_name}
                )
            config["groupMappings"][alias] = canonical_id
            # Persist the old ambiguous mapping only under an explicitly
            # disabled audit key. The backup below retains its original file.
            config.pop("groupTags", None)
            self._backup_private_config()
            self._atomic_json(self.config_path, config)
        return {
            "success": True,
            "rawToken": raw_token,
            "alias": alias,
            "groupId": canonical_id,
            "displayName": config["canonicalGroups"][canonical_id]["displayName"],
        }

    def _backup_private_config(self) -> None:
        if not self.config_path.exists():
            return
        stamp = datetime.now().strftime("%Y-%m-%d-%H%M%S-%f")
        shutil.copy2(self.config_path, self.backups_root / f"config-{stamp}.json")

    def analyze(
        self,
        stream: BinaryIO,
        filename: str,
        start_date: str,
        end_date: str,
        *,
        sheet_name: str = "",
        source_modified_at: str | None = None,
    ) -> dict[str, Any]:
        start = self._parse_requested_date(start_date, "開始日")
        end = self._parse_requested_date(end_date, "終了日")
        if end < start:
            raise TimetableError("終了日は開始日以降にしてください。")
        if (end - start).days > 400:
            raise TimetableError("一度に確認できる期間は400日までです。")
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise TimetableError("Excelファイル（.xlsx または .xlsm）を選択してください。")

        descriptor, temp_name = tempfile.mkstemp(prefix="timetable-", suffix=Path(filename).suffix, dir=self.runtime_root)
        os.close(descriptor)
        temp_path = Path(temp_name)
        try:
            with temp_path.open("wb") as handle:
                shutil.copyfileobj(stream, handle)
            config = self._config()
            config["_validCourseIds"] = {item["id"] for item in self._course_options()}
            parsed = self._parse_workbook(temp_path, sheet_name or config["sheetName"], config)
        except TimetableError:
            raise
        except Exception as error:
            raise TimetableError(
                "Excelを解析できませんでした。ファイルが破損していないか確認してください。",
                code="workbook_read_failed",
            ) from error
        finally:
            temp_path.unlink(missing_ok=True)

        entries = [item for item in parsed["entries"] if start <= date.fromisoformat(item["date"]) <= end]
        token_occurrences = [
            item for item in parsed.get("tokenOccurrences", [])
            if start <= date.fromisoformat(item["date"]) <= end
        ]
        if not entries:
            parsed["issues"].append(self._issue(
                "blocking", "no_entries_in_range", "指定期間内に授業・行事を見つけられませんでした。"
            ))
        previous = self._load_snapshot_entries()
        current_range_dates = {item["date"] for item in entries}
        old_in_range = [item for item in previous if start <= date.fromisoformat(item["date"]) <= end]
        changes = self.compare(old_in_range, entries)
        for change in changes:
            if change["type"] == "removed":
                before = change.get("before") or {}
                parsed["issues"].append(self._issue(
                    "warning",
                    "previous_entry_now_blank",
                    "前回登録されていた授業が今回は空白です。反映前に確認してください。",
                    entry=before,
                ))
        outside = [item for item in previous if not (start <= date.fromisoformat(item["date"]) <= end)]
        merged_snapshot = sorted(outside + entries, key=self._entry_sort_key)
        token = uuid4().hex
        preview = TimetablePreview(
            token=token,
            created_at=datetime.now(timezone.utc),
            filename=Path(filename).name,
            sheet_name=parsed["sheetName"],
            start_date=start,
            end_date=end,
            source_modified_at=source_modified_at,
            entries=entries,
            periods=parsed["periods"],
            issues=self._deduplicate_issues(parsed["issues"]),
            changes=changes,
            merged_snapshot=merged_snapshot,
            token_summary=self._token_summary(token_occurrences, config),
            token_occurrences=token_occurrences,
            comparison_base=old_in_range,
            date_count=0,
        )
        with self.lock:
            self._discard_old_previews()
            self.previews[token] = preview
        parsed_date_count = sum(
            start <= date.fromisoformat(value) <= end for value in parsed.get("dates", current_range_dates)
        )
        preview.date_count = parsed_date_count
        return self._preview_payload(preview)

    @staticmethod
    def _parse_requested_date(value: str, label: str) -> date:
        try:
            return date.fromisoformat(value)
        except (TypeError, ValueError) as error:
            raise TimetableError(f"{label}を正しい日付で指定してください。") from error

    def _parse_workbook(self, path: Path, sheet_name: str, config: dict[str, Any]) -> dict[str, Any]:
        try:
            formula_book = load_workbook(path, data_only=False, read_only=False)
            value_book = load_workbook(path, data_only=True, read_only=False)
        except Exception as error:
            raise TimetableError("Excelファイルを開けませんでした。", code="workbook_read_failed") from error
        if sheet_name not in formula_book.sheetnames:
            raise TimetableError(
                f"シート「{sheet_name}」が見つかりません。対象シートを確認してください。",
                code="sheet_not_found",
            )
        sheet = formula_book[sheet_name]
        value_sheet = value_book[sheet_name]
        issues: list[dict[str, Any]] = []
        # T is the Sunday subject column. U/V may legitimately be empty for the
        # whole workbook, in which case openpyxl does not include them in max_column.
        if sheet.max_column < 20:
            issues.append(self._issue(
                "blocking", "unexpected_structure", "曜日ごとの列（B～V列）を確認できません。"
            ))
        merged_values, merged_ranges = self._merged_values(sheet)
        entries: list[dict[str, Any]] = []
        periods_seen: dict[str, dict[str, str]] = {}
        active_dates: dict[int, date] = {}
        date_rows = 0
        parsed_dates: set[date] = set()
        formula_cache: dict[str, date | None] = {}
        active_period: int | None = None
        active_period_clocks: list[str] = []
        supplemental_zone = False
        seen_subject_origins: set[tuple[str, str, int]] = set()
        seen_schedule_subject_sources: set[tuple[str, str]] = set()
        see_below_dates: set[str] = set()
        token_occurrences: list[dict[str, Any]] = []

        def finish_period() -> None:
            nonlocal active_period, active_period_clocks, supplemental_zone
            if active_period is None:
                return
            if len(active_period_clocks) >= 2:
                timing = {"start": active_period_clocks[0], "end": active_period_clocks[-1]}
                key = str(active_period)
                if key in periods_seen and periods_seen[key] != timing:
                    issues.append(self._issue(
                        "blocking",
                        "inconsistent_period_time",
                        f"{active_period}限の時刻が週によって異なります。",
                    ))
                periods_seen[key] = timing
            finished_period = active_period
            active_period = None
            active_period_clocks = []
            supplemental_zone = finished_period == 4

        for row in range(1, sheet.max_row + 1):
            found_dates: dict[int, date] = {}
            for day_index, column in enumerate(DAY_COLUMNS):
                parsed_date = self._cell_date(
                    sheet, value_sheet, row, column, formula_cache, set()
                )
                if parsed_date:
                    found_dates[day_index] = parsed_date
            if found_dates:
                finish_period()
                supplemental_zone = False
                date_rows += 1
                parsed_dates.update(found_dates.values())
                active_dates = found_dates
                for day_index, parsed_date in found_dates.items():
                    if parsed_date.weekday() != day_index:
                        issues.append(self._issue(
                            "warning",
                            "weekday_mismatch",
                            f"{parsed_date.isoformat()} の曜日列が一致しません。Excelの見出しを確認してください。",
                            date_value=parsed_date,
                        ))
                if len(found_dates) != 7:
                    issues.append(self._issue(
                        "blocking", "date_parse_failed", "週の日付を7日分解析できませんでした。",
                        date_value=next(iter(found_dates.values())),
                    ))
                continue

            period_value = self._logical_value(sheet, row, 1, merged_values)
            period_text = _text(period_value)
            parsed_period, inline_timing = self._parse_period(period_text)
            if parsed_period is not None:
                finish_period()
                supplemental_zone = False
                active_period = parsed_period
                if inline_timing:
                    active_period_clocks = [inline_timing["start"], inline_timing["end"]]
            elif active_period is not None:
                clock = self._parse_clock(period_value)
                if clock and clock not in active_period_clocks:
                    active_period_clocks.append(clock)

            if not active_dates:
                continue
            for day_index, subject_column in enumerate(DAY_COLUMNS):
                subject_value = self._logical_value(sheet, row, subject_column, merged_values)
                if self._parse_date_value(subject_value, sheet.parent.epoch):
                    continue
                # Keep the exact visible source spelling for staff review and
                # normalize only inside parse_subject().
                subject_raw = _raw_text(subject_value)
                if not subject_raw:
                    continue
                grade_value = self._logical_value(sheet, row, subject_column + 1, merged_values)
                room_value = self._logical_value(sheet, row, subject_column + 2, merged_values)
                grade_raw = _raw_text(grade_value)
                room_raw = _raw_text(room_value)
                source_cell = sheet.cell(row=row, column=subject_column).coordinate
                source = {"sourceRow": row, "sourceCell": source_cell}
                if self._is_column_header(subject_raw, grade_raw, room_raw):
                    continue
                if day_index not in active_dates:
                    issues.append(self._issue(
                        "blocking", "date_parse_failed", "授業セルに対応する日付を解析できませんでした。",
                        extra=source,
                    ))
                    continue
                entry_date = active_dates[day_index]
                subject_merge = merged_ranges.get((row, subject_column))
                origin = subject_merge or source_cell
                if active_period is None:
                    if supplemental_zone:
                        if (origin, entry_date.isoformat()) in seen_schedule_subject_sources:
                            continue
                        category = self._supplemental_category(
                            subject_raw, grade_raw, room_raw,
                            linked_to_see_below=entry_date.isoformat() in see_below_dates,
                        )
                        issues.append(self._issue(
                            "reference",
                            "supplemental_info",
                            self._supplemental_message(category),
                            date_value=entry_date,
                            subject_raw=subject_raw,
                            extra={**source, "referenceCategory": category},
                        ))
                    elif "下記参照" in subject_raw:
                        issues.append(self._issue(
                            "warning", "see_below", "「下記参照」があります。Excel下部の注記を確認してください。",
                            date_value=entry_date, subject_raw=subject_raw, extra=source,
                        ))
                    else:
                        issues.append(self._issue(
                            "blocking", "period_parse_failed", "授業セルに対応する時限を特定できませんでした。",
                            date_value=entry_date, subject_raw=subject_raw, extra=source,
                        ))
                    continue

                origin_key = (origin, entry_date.isoformat(), active_period)
                if origin_key in seen_subject_origins:
                    continue
                seen_subject_origins.add(origin_key)
                seen_schedule_subject_sources.add((origin, entry_date.isoformat()))
                room_cell = sheet.cell(row=row, column=subject_column + 2)
                room_merge = merged_ranges.get((row, subject_column + 2))
                room_number_format = (
                    sheet[room_merge.split(":", 1)[0]].number_format
                    if room_merge
                    else room_cell.number_format
                )
                entry, entry_issues, subject_tokens = self._normalize_entry(
                    entry_date, active_period, subject_raw, grade_value,
                    room_value, room_number_format, config,
                )
                entry["_sourceRow"] = source.get("sourceRow")
                entry["_sourceCell"] = source.get("sourceCell")
                entries.append(entry)
                for token in subject_tokens:
                    token_occurrences.append({
                        **token,
                        "date": entry["date"],
                        "period": entry["period"],
                        "subjectRaw": entry["subjectRaw"],
                        "subjectName": entry["subjectName"],
                        "grades": entry["grades"],
                        "gradeRaw": entry["gradeRaw"],
                        **source,
                    })
                for issue in entry_issues:
                    issue.update(source)
                issues.extend(entry_issues)
                if "下記参照" in room_raw:
                    see_below_dates.add(entry["date"])
                    issues.append(self._issue(
                        "warning", "see_below", "教室欄に「下記参照」があります。Excel下部の注記を確認してください。",
                        entry=entry, extra=source,
                    ))
                for column, field in ((subject_column + 1, "学年"), (subject_column + 2, "教室")):
                    merge = merged_ranges.get((row, column))
                    if merge and self._logical_value(sheet, row, column, merged_values) is None:
                        issues.append(self._issue(
                            "warning", "merged_value_unavailable", f"結合セルから{field}を復元できませんでした。",
                            entry=entry, extra=source,
                        ))

            if active_period is not None and sheet.cell(row=row, column=1).border.bottom.style:
                finish_period()

        finish_period()

        if date_rows == 0:
            issues.append(self._issue("blocking", "date_parse_failed", "Excel内の日付を解析できませんでした。"))
        if not entries:
            issues.append(self._issue("blocking", "unexpected_structure", "時間割の授業行を解析できませんでした。"))
        if set(periods_seen) != {"1", "2", "3", "4"}:
            missing = sorted({"1", "2", "3", "4"} - set(periods_seen))
            issues.append(self._issue(
                "warning", "period_times_missing", f"{', '.join(missing)}限の時刻をExcelから取得できないため標準時刻を使用します。"
            ))
        periods = copy.deepcopy(DEFAULT_PERIODS)
        periods.update(periods_seen)
        self._detect_duplicates(entries, issues)
        return {
            "sheetName": sheet_name,
            "entries": sorted(entries, key=self._entry_sort_key),
            "periods": periods,
            "issues": issues,
            "dates": [value.isoformat() for value in sorted(parsed_dates)],
            "tokenOccurrences": token_occurrences,
        }

    @staticmethod
    def _merged_values(sheet: Any) -> tuple[dict[tuple[int, int], Any], dict[tuple[int, int], str]]:
        values: dict[tuple[int, int], Any] = {}
        ranges: dict[tuple[int, int], str] = {}
        for merged in sheet.merged_cells.ranges:
            anchor = sheet.cell(merged.min_row, merged.min_col).value
            for row in range(merged.min_row, merged.max_row + 1):
                for column in range(merged.min_col, merged.max_col + 1):
                    values[(row, column)] = anchor
                    ranges[(row, column)] = str(merged)
        return values, ranges

    @staticmethod
    def _logical_value(sheet: Any, row: int, column: int, merged_values: dict[tuple[int, int], Any]) -> Any:
        if (row, column) in merged_values:
            return merged_values[(row, column)]
        cell = sheet.cell(row=row, column=column)
        return None if isinstance(cell, MergedCell) else cell.value

    def _cell_date(
        self, sheet: Any, value_sheet: Any, row: int, column: int,
        cache: dict[str, date | None], visiting: set[str],
    ) -> date | None:
        coordinate = sheet.cell(row=row, column=column).coordinate
        if coordinate in cache:
            return cache[coordinate]
        cached_value = value_sheet.cell(row=row, column=column).value
        parsed = self._parse_date_value(cached_value, sheet.parent.epoch)
        if parsed:
            cache[coordinate] = parsed
            return parsed
        value = sheet.cell(row=row, column=column).value
        parsed = self._parse_date_value(value, sheet.parent.epoch)
        if parsed:
            cache[coordinate] = parsed
            return parsed
        if isinstance(value, str) and value.startswith("=") and coordinate not in visiting:
            visiting = set(visiting) | {coordinate}
            formula = value[1:].replace("$", "").strip()
            date_match = re.fullmatch(r"DATE\((\d{4}),\s*(\d{1,2}),\s*(\d{1,2})\)", formula, re.I)
            if date_match:
                try:
                    parsed = date(*(int(part) for part in date_match.groups()))
                except ValueError:
                    parsed = None
            else:
                reference = re.fullmatch(r"(?:'[^']+'|[^!]+)!([A-Z]+\d+)(?:\s*\+\s*(\d+))?|([A-Z]+\d+)(?:\s*\+\s*(\d+))?", formula, re.I)
                if reference:
                    ref = reference.group(1) or reference.group(3)
                    offset = int(reference.group(2) or reference.group(4) or 0)
                    letters, ref_row = coordinate_from_string(ref.upper())
                    base = self._cell_date(
                        sheet, value_sheet, ref_row, column_index_from_string(letters), cache, visiting
                    )
                    parsed = base + timedelta(days=offset) if base else None
        cache[coordinate] = parsed
        return parsed

    @staticmethod
    def _parse_date_value(value: Any, epoch: datetime) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and 30000 <= value <= 80000:
            try:
                return from_excel(value, epoch).date()
            except (TypeError, ValueError, OverflowError):
                return None
        if not isinstance(value, str) or value.startswith("="):
            return None
        candidate = unicodedata.normalize("NFKC", value).strip()
        match = re.fullmatch(r"(20\d{2})[./年-](\d{1,2})[./月-](\d{1,2})日?", candidate)
        if not match:
            return None
        try:
            return date(*(int(part) for part in match.groups()))
        except ValueError:
            return None

    @staticmethod
    def _parse_period(value: str) -> tuple[int | None, dict[str, str] | None]:
        if not value:
            return None, None
        normalized = unicodedata.normalize("NFKC", value).strip()
        period_match = re.match(r"^([1-4])\s*(?:限|時限)?(?:\s|$)", normalized)
        if not period_match:
            return None, None
        period = int(period_match.group(1))
        times = re.findall(r"(?:[01]?\d|2[0-3]):[0-5]\d", normalized)
        timing = {"start": times[0].zfill(5), "end": times[1].zfill(5)} if len(times) >= 2 else None
        return period, timing

    @staticmethod
    def _parse_clock(value: Any) -> str | None:
        if isinstance(value, time):
            return value.strftime("%H:%M")
        text = _text(value)
        match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)(?::[0-5]\d)?", text)
        if not match:
            return None
        return f"{int(match.group(1)):02d}:{match.group(2)}"

    @staticmethod
    def _is_column_header(subject: str, grade: str, room: str) -> bool:
        normalized_subject = _mapping_key(subject)
        return (
            "科目" in normalized_subject
            and "講師" in normalized_subject
            and _mapping_key(grade) == "学年"
            and _mapping_key(room) == "教室"
        )

    @staticmethod
    def _supplemental_category(
        subject: str, grade: str, room: str, *, linked_to_see_below: bool
    ) -> str:
        if linked_to_see_below or "下記参照" in subject or "下記参照" in room:
            return "see_below_detail"
        normalized = unicodedata.normalize("NFKC", subject)
        if grade or room or re.search(r"(?:組|クラス)\s*[:：]|[↓→]", normalized):
            return "schedule_detail"
        if re.search(r"(?:予定|候補|説明会|講話|研修|行事|イベント)", normalized):
            return "special_schedule"
        return "note"

    @staticmethod
    def _supplemental_message(category: str) -> str:
        messages = {
            "see_below_detail": "「下記参照」に関連する可能性がある補足記載です。",
            "schedule_detail": "クラス・教室などの時間割補足が記載されています。",
            "special_schedule": "予定・行事に関する補足が記載されています。",
            "note": "Excelの補足欄に参考情報が記載されています。",
        }
        return messages[category]

    @staticmethod
    def _resolve_subject_mapping(
        parsed_subject: dict[str, Any], config: dict[str, Any], valid_course_ids: set[str]
    ) -> dict[str, Any]:
        mapping_name = parsed_subject["mappingName"]
        key = _mapping_key(mapping_name)
        item_mapping = config["timetableItemMappings"].get(key) if key else None
        if item_mapping:
            return {
                "kind": item_mapping["type"],
                "courseId": None,
                "missingCourseId": None,
                "explicitItem": True,
            }
        kind = parsed_subject["kind"]
        course_id = None
        if kind in {"class", "exam"} and mapping_name:
            course_id = config["subjectMappings"].get(key)
        if course_id and course_id not in valid_course_ids:
            return {
                "kind": kind,
                "courseId": None,
                "missingCourseId": course_id,
                "explicitItem": False,
            }
        return {
            "kind": kind,
            "courseId": course_id,
            "missingCourseId": None,
            "explicitItem": False,
        }

    def _normalize_entry(
        self, entry_date: date, period: int, subject_raw: str, grade_value: Any,
        room_value: Any, number_format: str, config: dict[str, Any],
    ) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, str]]]:
        issues: list[dict[str, Any]] = []
        parsed_subject = self.parse_subject(subject_raw)
        grades, grade_raw = self.parse_grades(grade_value)
        rooms, room_warning = self.parse_rooms(room_value, number_format)
        status = parsed_subject["status"]
        mapping_name = parsed_subject["mappingName"]
        resolution = self._resolve_subject_mapping(
            parsed_subject,
            config,
            config.get("_validCourseIds", set(config["subjectMappings"].values())),
        )
        kind = resolution["kind"]
        course_id = resolution["courseId"]
        missing_course_id = resolution["missingCourseId"]
        resolved_group_tags: list[str] = []
        unresolved_group_tokens: list[dict[str, str]] = []
        for token in parsed_subject["tokens"]:
            if token["tokenType"] != "group":
                continue
            alias = self._group_alias(token["rawToken"])
            canonical_id = config["groupMappings"].get(alias)
            canonical = config["canonicalGroups"].get(canonical_id, {}) if canonical_id else {}
            display_name = _text(canonical.get("displayName")) if isinstance(canonical, dict) else ""
            if display_name:
                resolved_group_tags.append(display_name)
            else:
                resolved_group_tags.append(token["tokenValue"])
                unresolved_group_tokens.append({**token, "alias": alias})
        entry = {
            "date": entry_date.isoformat(),
            "period": period,
            "courseId": course_id,
            "subjectRaw": subject_raw,
            "subjectName": parsed_subject["subjectName"],
            "subjectBaseName": parsed_subject["subjectBaseName"],
            "courseVariantTags": parsed_subject["courseVariantTags"],
            "grades": grades,
            "gradeRaw": grade_raw,
            "groupTags": list(dict.fromkeys(resolved_group_tags)),
            "instructors": parsed_subject["instructors"],
            "rooms": rooms,
            "kind": kind,
            "status": status,
        }
        entry["entryId"] = self._entry_id(entry)
        if grade_raw and grades is None:
            issues.append(self._issue("blocking", "invalid_grade", f"学年「{grade_raw}」を解釈できません。", entry=entry))
            entry["grades"] = []
        elif not grade_raw:
            entry["grades"] = []
        if missing_course_id:
            issues.append(self._issue(
                "blocking", "mapped_course_missing",
                f"「{mapping_name}」の対応先授業が見つかりません: {missing_course_id}",
                entry=entry,
                extra={"subjectName": mapping_name, "courseId": missing_course_id},
            ))
        elif (
            not resolution["explicitItem"]
            and kind in {"class", "exam"}
            and mapping_name
            and course_id is None
        ):
            issues.append(self._issue(
                "blocking", "unregistered_subject", f"未登録の科目があります: {mapping_name}", entry=entry,
                extra={"subjectName": mapping_name},
            ))
        if kind == "exam" and not mapping_name:
            issues.append(self._issue("blocking", "empty_exam_name", "試験名が空です。", entry=entry))
        for token in unresolved_group_tokens:
            issues.append(self._issue(
                "blocking", "unknown_group_tag",
                f"未登録の受講グループ表記があります: {token['rawToken']}", entry=entry,
                extra={
                    "groupTag": token["tokenValue"],
                    "rawGroupToken": token["rawToken"],
                    "groupAlias": token["alias"],
                },
            ))
        if status == "tentative":
            issues.append(self._issue("warning", "tentative", "候補日または予定として記載されています。", entry=entry))
        if "下記参照" in subject_raw:
            issues.append(self._issue("warning", "see_below", "「下記参照」があります。Excel下部の注記を確認してください。", entry=entry))
        if room_warning:
            issues.append(self._issue("warning", "suspicious_room", room_warning, entry=entry))
        return entry, issues, parsed_subject["tokens"]

    @staticmethod
    def parse_subject(subject_raw: str) -> dict[str, Any]:
        raw_value = subject_raw.strip()
        tokens: list[dict[str, str]] = []
        removable_spans: list[tuple[int, int]] = []
        variant_spans: list[tuple[int, int]] = []
        status = "confirmed"
        instructors: list[str] = []
        for match in re.finditer(r"【([^【】]*)】", raw_value):
            instructor_raw = _text(match.group(1))
            if instructor_raw:
                split = [part.strip() for part in re.split(r"[,、/・]+", instructor_raw) if part.strip()]
                instructors.extend(split or [instructor_raw])
            tokens.append({
                "rawToken": match.group(0), "tokenType": "instructor", "tokenValue": instructor_raw,
            })
            removable_spans.append(match.span())

        course_variant_tags: list[str] = []
        for match in re.finditer(r"\[([^\[\]]+)\]", raw_value):
            token_value = _text(match.group(1))
            if token_value == "予定":
                status = "tentative"
                token_type = "status"
                removable_spans.append(match.span())
            else:
                token_type = "course-variant"
                course_variant_tags.append(token_value)
                variant_spans.append(match.span())
            tokens.append({
                "rawToken": match.group(0), "tokenType": token_type, "tokenValue": token_value,
            })

        group_tags: list[str] = []
        for match in re.finditer(r"（([^（）]+)）|\(([^()]+)\)", raw_value):
            token_value = _text(match.group(1) or match.group(2))
            compact = re.sub(r"\s+", "", token_value)
            if compact == "候補日":
                status = "tentative"
                token_type = "status"
                removable_spans.append(match.span())
            elif TimetableService._is_group_token_value(compact):
                token_type = "group"
                group_tags.append(compact)
                removable_spans.append(match.span())
            else:
                token_type = "subject-qualifier"
            tokens.append({
                "rawToken": match.group(0), "tokenType": token_type, "tokenValue": token_value,
            })

        def without(spans: list[tuple[int, int]]) -> str:
            characters = list(raw_value)
            for start, end in spans:
                characters[start:end] = " " * (end - start)
            normalized = re.sub(
                r"\s+", " ", unicodedata.normalize("NFKC", "".join(characters))
            ).strip()
            return re.sub(r"\s+(?=\[)", "", normalized)

        subject_name = without(removable_spans)
        subject_base_name = without(removable_spans + variant_spans)
        kind = "class"
        if "下記参照" in subject_name:
            kind = "other"
        elif re.match(r"^試験\s*[:：]", subject_name):
            kind = "exam"
        elif re.search(r"(?:休暇|休業|休日|休校)", subject_name):
            kind = "holiday"
        elif re.search(r"(?:説明会|発表会|講話|特別行事|オリエンテーション|入学式|卒業式)", subject_name):
            kind = "event"
        mapping_name = re.sub(r"^試験\s*[:：]\s*", "", subject_name).strip() if kind == "exam" else subject_name
        mapping_base_name = (
            re.sub(r"^試験\s*[:：]\s*", "", subject_base_name).strip()
            if kind == "exam" else subject_base_name
        )
        return {
            "subjectName": subject_name,
            "subjectBaseName": subject_base_name,
            "mappingName": mapping_name,
            "mappingBaseName": mapping_base_name,
            "courseVariantTags": list(dict.fromkeys(course_variant_tags)),
            "groupTags": list(dict.fromkeys(group_tags)),
            "instructors": list(dict.fromkeys(instructors)),
            "tokens": tokens,
            "kind": kind,
            "status": status,
        }

    @staticmethod
    def _is_group_token_value(value: str) -> bool:
        """Return true only for conservative class/group-like round tokens."""
        return bool(re.fullmatch(
            r"(?:"
            r"(?:[A-Z](?:[/・][A-Z])*)(?:組|クラス)?"
            r"|(?:[1-9](?:[/・][1-9])*)(?:組|クラス)"
            r"|全(?:組|クラス)"
            r")",
            value,
            re.IGNORECASE,
        ))

    @staticmethod
    def parse_grades(value: Any) -> tuple[list[int] | None, str]:
        raw = _raw_text(value)
        if not raw:
            return [], ""
        compact = re.sub(r"\s+|年生?|学年", "", unicodedata.normalize("NFKC", raw))
        if compact in {"全", "全学年"}:
            return [1, 2, 3], raw
        parts = [part for part in re.split(r"[,、/&・]+", compact) if part]
        if parts and all(part in {"1", "2", "3"} for part in parts):
            return sorted({int(part) for part in parts}), raw
        return None, raw

    @staticmethod
    def parse_rooms(value: Any, number_format: str = "General") -> tuple[list[str], str | None]:
        if value is None or _text(value) == "":
            return [], None
        if isinstance(value, float) and not value.is_integer():
            text = str(value)
            return [text], f"教室表記「{text}」に小数点があります。元Excelを確認してください。"
        if isinstance(value, (int, float)):
            integer = int(value)
            fmt = number_format or "General"
            if "," in fmt and len(str(abs(integer))) > 3:
                text = f"{integer:,}"
            else:
                text = str(integer)
        else:
            text = _text(value)
        if re.fullmatch(r"\d+\.\d+", text):
            return [text], f"教室表記「{text}」に小数点があります。元Excelを確認してください。"
        rooms = [part.strip() for part in re.split(r"[,、・/]+", text) if part.strip()]
        return rooms or [text], None

    def _token_summary(
        self, occurrences: list[dict[str, Any]], config: dict[str, Any]
    ) -> dict[str, list[dict[str, Any]]]:
        category_names = {
            "course-variant": "courseVariants",
            "group": "groups",
            "instructor": "instructors",
            "status": "statusMarkers",
            "subject-qualifier": "subjectQualifiers",
        }
        buckets: dict[str, dict[str, list[dict[str, Any]]]] = {
            category: defaultdict(list) for category in category_names.values()
        }
        for occurrence in occurrences:
            category = category_names.get(occurrence.get("tokenType"))
            if not category:
                continue
            if occurrence["tokenType"] == "group":
                key = self._group_alias(occurrence["rawToken"])
            else:
                key = unicodedata.normalize("NFKC", occurrence["rawToken"])
            buckets[category][key].append(occurrence)

        summary: dict[str, list[dict[str, Any]]] = {}
        for category, groups in buckets.items():
            items = []
            for key, uses in groups.items():
                raw_counts = Counter(item["rawToken"] for item in uses)
                item: dict[str, Any] = {
                    "key": key,
                    "tokenValue": uses[0]["tokenValue"],
                    "count": len(uses),
                    "rawTokens": [
                        {"rawToken": raw_token, "count": count}
                        for raw_token, count in sorted(raw_counts.items())
                    ],
                    "examples": [
                        {
                            field: use.get(field)
                            for field in (
                                "rawToken", "subjectRaw", "subjectName", "date", "period",
                                "grades", "gradeRaw", "sourceCell", "sourceRow",
                            )
                        }
                        for use in uses
                    ],
                }
                if category == "groups":
                    canonical_id = config["groupMappings"].get(key)
                    canonical = config["canonicalGroups"].get(canonical_id, {}) if canonical_id else {}
                    display_name = _text(canonical.get("displayName")) if isinstance(canonical, dict) else ""
                    item["resolved"] = bool(display_name)
                    item["canonicalGroup"] = (
                        {"id": canonical_id, "displayName": display_name} if display_name else None
                    )
                items.append(item)
            summary[category] = sorted(items, key=lambda item: (-item["count"], item["key"]))
        return summary

    @staticmethod
    def _entry_id(entry: dict[str, Any]) -> str:
        identity = {
            "date": entry["date"], "period": entry["period"],
            "courseId": entry.get("courseId"), "subjectName": entry["subjectName"],
            "subjectBaseName": entry.get("subjectBaseName"),
            "courseVariantTags": entry.get("courseVariantTags", []),
            "grades": entry.get("grades", []), "groupTags": entry.get("groupTags", []),
        }
        digest = hashlib.sha256(json.dumps(identity, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:16]
        return f"tt-{digest}"

    def _detect_duplicates(self, entries: list[dict[str, Any]], issues: list[dict[str, Any]]) -> None:
        groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
        for entry in entries:
            key = (
                entry["date"], entry["period"], entry.get("courseId"), entry["subjectName"],
                tuple(entry.get("grades", [])), tuple(entry.get("groupTags", [])),
                tuple(entry.get("instructors", [])), tuple(entry.get("rooms", [])),
                entry.get("kind"), entry.get("status"),
            )
            groups[key].append(entry)
        for duplicates in groups.values():
            if len(duplicates) > 1:
                issues.append(self._issue(
                    "warning", "duplicate_entry", "同じ対象の授業が同一日時・時限に重複しています。",
                    entry=duplicates[0],
                ))

    @staticmethod
    def _issue(
        severity: str, code: str, message: str, *, entry: dict[str, Any] | None = None,
        date_value: date | None = None, subject_raw: str = "", extra: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        item: dict[str, Any] = {"severity": severity, "code": code, "message": message}
        if entry:
            item.update({
                "date": entry.get("date"), "period": entry.get("period"),
                "subjectRaw": entry.get("subjectRaw", ""),
            })
        elif date_value:
            item["date"] = date_value.isoformat()
            if subject_raw:
                item["subjectRaw"] = subject_raw
        if extra:
            item.update(extra)
        item["id"] = hashlib.sha256(json.dumps(item, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:16]
        return item

    @staticmethod
    def _deduplicate_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen: set[tuple[Any, ...]] = set()
        result = []
        for issue in issues:
            if issue["code"] in {"unregistered_subject", "mapped_course_missing", "unknown_group_tag"}:
                key = (
                    issue["code"], issue.get("subjectName"),
                    issue.get("groupAlias") or issue.get("groupTag"),
                )
            else:
                key = (issue["code"], issue.get("date"), issue.get("period"), issue.get("subjectRaw"), issue["message"])
            if key not in seen:
                seen.add(key)
                result.append(issue)
        return result

    def _load_snapshot_entries(self) -> list[dict[str, Any]]:
        document = self._load_json(self.snapshot_path, {"entries": []})
        entries = document.get("entries", []) if isinstance(document, dict) else []
        if not isinstance(entries, list):
            raise TimetableError("前回の時間割スナップショットが不正です。", status=500)
        upgraded = []
        for source_entry in entries:
            entry = copy.deepcopy(source_entry)
            if "subjectBaseName" not in entry or "courseVariantTags" not in entry:
                parsed = self.parse_subject(_text(entry.get("subjectRaw") or entry.get("subjectName")))
                entry.setdefault("subjectBaseName", parsed["subjectBaseName"])
                entry.setdefault("courseVariantTags", parsed["courseVariantTags"])
            upgraded.append(entry)
        return upgraded

    @classmethod
    def compare(cls, before: list[dict[str, Any]], after: list[dict[str, Any]]) -> list[dict[str, Any]]:
        old_slots: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
        new_slots: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
        for entry in before:
            old_slots[(entry["date"], entry["period"])].append(entry)
        for entry in after:
            new_slots[(entry["date"], entry["period"])].append(entry)
        changes: list[dict[str, Any]] = []
        for slot in sorted(set(old_slots) | set(new_slots)):
            old = list(old_slots.get(slot, []))
            new = list(new_slots.get(slot, []))
            pairs: list[tuple[dict[str, Any], dict[str, Any]]] = []
            used_new: set[int] = set()
            unmatched_old = []
            for old_entry in old:
                identity = cls._comparison_identity(old_entry)
                matches = [i for i, new_entry in enumerate(new) if i not in used_new and cls._comparison_identity(new_entry) == identity]
                if len(matches) == 1:
                    index = matches[0]
                    used_new.add(index)
                    pairs.append((old_entry, new[index]))
                else:
                    unmatched_old.append(old_entry)
            unmatched_new = [entry for index, entry in enumerate(new) if index not in used_new]
            fallback_old: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
            fallback_new: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
            for entry in unmatched_old:
                fallback_old[cls._target_identity(entry)].append(entry)
            for entry in unmatched_new:
                fallback_new[cls._target_identity(entry)].append(entry)
            consumed_old: set[int] = set()
            consumed_new: set[int] = set()
            for target in set(fallback_old) & set(fallback_new):
                if len(fallback_old[target]) == 1 and len(fallback_new[target]) == 1:
                    old_entry, new_entry = fallback_old[target][0], fallback_new[target][0]
                    pairs.append((old_entry, new_entry))
                    consumed_old.add(id(old_entry))
                    consumed_new.add(id(new_entry))
            for old_entry, new_entry in pairs:
                fields = cls._changed_fields(old_entry, new_entry)
                if fields:
                    changes.append({"type": "changed", "date": slot[0], "period": slot[1], "fields": fields, "before": old_entry, "after": new_entry})
            for entry in unmatched_old:
                if id(entry) not in consumed_old:
                    changes.append({"type": "removed", "date": slot[0], "period": slot[1], "fields": [], "before": entry, "after": None})
            for entry in unmatched_new:
                if id(entry) not in consumed_new:
                    changes.append({"type": "added", "date": slot[0], "period": slot[1], "fields": [], "before": None, "after": entry})
        return changes

    @staticmethod
    def _comparison_identity(entry: dict[str, Any]) -> tuple[Any, ...]:
        return (
            entry.get("courseId"), entry.get("subjectName"),
            tuple(entry.get("grades", [])), tuple(entry.get("groupTags", [])),
        )

    @staticmethod
    def _target_identity(entry: dict[str, Any]) -> tuple[Any, ...]:
        return (tuple(entry.get("grades", [])), tuple(entry.get("groupTags", [])))

    @staticmethod
    def _changed_fields(before: dict[str, Any], after: dict[str, Any]) -> list[str]:
        labels = {
            "courseId": "科目", "subjectName": "科目", "rooms": "教室", "instructors": "講師",
            "grades": "学年", "groupTags": "グループ", "kind": "授業種別", "status": "確定状態",
        }
        return list(dict.fromkeys(labels[field] for field in labels if before.get(field) != after.get(field)))

    @staticmethod
    def _entry_sort_key(entry: dict[str, Any]) -> tuple[Any, ...]:
        return (
            entry["date"], entry["period"], tuple(entry.get("grades", [])),
            tuple(entry.get("groupTags", [])), entry.get("subjectName", ""),
        )

    def _subject_usage_summary(
        self, preview: TimetablePreview, subject_name: str
    ) -> dict[str, Any] | None:
        key = _mapping_key(subject_name)
        uses = []
        parsed_example = None
        for entry in preview.entries:
            parsed = self.parse_subject(entry.get("subjectRaw", ""))
            if _mapping_key(parsed["mappingName"]) != key:
                continue
            parsed_example = parsed_example or parsed
            uses.append(entry)
        if not uses or parsed_example is None:
            return None
        options = self._course_options()
        instructors = sorted({
            instructor
            for entry in uses for instructor in entry.get("instructors", [])
            if instructor
        })
        years = Counter(entry["date"][:4] for entry in uses if entry.get("date"))
        academic_year = years.most_common(1)[0][0] if years else None
        exact = [item for item in options if item["title"] == subject_name]
        base_name = parsed_example["mappingBaseName"]
        related = [
            item for item in options
            if item["title"] == base_name and item["title"] != subject_name
        ]
        return {
            "subjectName": subject_name,
            "subjectBaseName": base_name,
            "courseVariantTags": parsed_example["courseVariantTags"],
            "count": len(uses),
            "academicYearCandidate": academic_year,
            "instructorCandidates": instructors,
            "exactCourseCandidates": exact,
            "relatedBaseCourses": related,
            "examples": [
                {
                    "date": entry.get("date"),
                    "period": entry.get("period"),
                    "grades": entry.get("grades", []),
                    "gradeRaw": entry.get("gradeRaw", ""),
                    "groupTags": entry.get("groupTags", []),
                    "instructors": entry.get("instructors", []),
                    "rooms": entry.get("rooms", []),
                    "subjectRaw": entry.get("subjectRaw", ""),
                    "sourceRow": entry.get("_sourceRow"),
                    "sourceCell": entry.get("_sourceCell"),
                }
                for entry in uses
            ],
        }

    def _subject_summary(self, preview: TimetablePreview) -> list[dict[str, Any]]:
        unresolved = {
            issue.get("subjectName")
            for issue in preview.issues
            if issue.get("code") == "unregistered_subject" and issue.get("subjectName")
        }
        summaries = [
            summary
            for subject_name in sorted(unresolved)
            if (summary := self._subject_usage_summary(preview, subject_name)) is not None
        ]
        return sorted(
            summaries,
            key=lambda item: (item["subjectBaseName"], item["subjectName"]),
        )

    def _mapping_summary(self, preview: TimetablePreview) -> list[dict[str, Any]]:
        config = self._config()
        courses = {item["id"]: item for item in self._course_options()}
        results = []
        for key in sorted(set(config["subjectMappings"]) | set(config["timetableItemMappings"])):
            item_mapping = config["timetableItemMappings"].get(key)
            display_name = _text(item_mapping.get("displayName")) if item_mapping else ""
            if not display_name:
                for entry in preview.entries:
                    parsed = self.parse_subject(entry.get("subjectRaw", ""))
                    if _mapping_key(parsed["mappingName"]) == key:
                        display_name = parsed["mappingName"]
                        break
            display_name = display_name or key
            usage = self._subject_usage_summary(preview, display_name) or {
                "count": 0,
                "examples": [],
            }
            course_id = config["subjectMappings"].get(key)
            kind = item_mapping["type"] if item_mapping else "class"
            course = courses.get(course_id) if course_id else None
            results.append({
                "subjectName": display_name,
                "classification": kind,
                "classificationLabel": KIND_LABELS[kind],
                "courseId": course_id,
                "courseTitle": course["title"] if course else None,
                "count": usage["count"],
                "examples": usage["examples"],
            })
        return results

    def _preview_payload(self, preview: TimetablePreview) -> dict[str, Any]:
        severity_counts = Counter(issue["severity"] for issue in preview.issues)
        grouped_counts: Counter[tuple[str, str, str]] = Counter()
        fallback_labels: dict[str, str] = {}
        for issue in preview.issues:
            category = issue.get("referenceCategory", "")
            grouped_counts[(issue["severity"], issue["code"], category)] += 1
            fallback_labels.setdefault(issue["code"], issue["message"])
        issue_counts = []
        for (severity, code, category), count in sorted(
            grouped_counts.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])
        ):
            issue_counts.append({
                "severity": severity,
                "code": code,
                "category": category or None,
                "label": ISSUE_LABELS.get(code, fallback_labels[code]),
                "count": count,
            })
        return {
            "token": preview.token,
            "filename": preview.filename,
            "sheetName": preview.sheet_name,
            "startDate": preview.start_date.isoformat(),
            "endDate": preview.end_date.isoformat(),
            "dateCount": preview.date_count,
            "entryCount": len(preview.entries),
            "warningCount": severity_counts["warning"],
            "blockingCount": severity_counts["blocking"],
            "referenceCount": severity_counts["reference"],
            "infoCount": severity_counts["info"],
            "issueCounts": issue_counts,
            "tokenSummary": preview.token_summary,
            "unregisteredSubjects": self._subject_summary(preview),
            "mappingSettings": self._mapping_summary(preview),
            "issues": preview.issues,
            "changes": preview.changes,
            "changeCounts": {
                "added": sum(item["type"] == "added" for item in preview.changes),
                "removed": sum(item["type"] == "removed" for item in preview.changes),
                "changed": sum(item["type"] == "changed" for item in preview.changes),
            },
        }

    def _discard_old_previews(self) -> None:
        cutoff = datetime.now(timezone.utc) - timedelta(hours=2)
        self.previews = {token: item for token, item in self.previews.items() if item.created_at >= cutoff}
        self.course_contexts = {
            token: item
            for token, item in self.course_contexts.items()
            if item.created_at >= cutoff and item.preview_token in self.previews
        }

    def get_preview(self, token: str) -> dict[str, Any]:
        with self.lock:
            self._discard_old_previews()
            preview = self.previews.get(token)
            if preview is None:
                raise TimetableError(
                    "確認内容の有効期限が切れました。Excelをもう一度読み込んでください。",
                    code="preview_expired", status=404,
                )
            return self._preview_payload(preview)

    def create_course_context(self, preview_token: str, subject_name: str) -> dict[str, Any]:
        with self.lock:
            self.get_preview(preview_token)
            preview = self.previews[preview_token]
            match = self._subject_usage_summary(preview, subject_name)
            if match is None:
                raise TimetableError(
                    "この項目は現在の時間割にありません。時間割を確認し直してください。",
                    code="subject_context_invalid", status=409,
                )
            key = _mapping_key(subject_name)
            if key in self._config()["subjectMappings"]:
                raise TimetableError(
                    "この科目はすでにClassViewの授業と対応済みです。",
                    code="subject_already_mapped", status=409,
                )
            token = uuid4().hex
            self.course_contexts[token] = PendingCourseContext(
                token=token,
                preview_token=preview_token,
                subject_name=subject_name,
                created_at=datetime.now(timezone.utc),
            )
            return {"token": token, "previewToken": preview_token, **match}

    def get_course_context(self, token: str) -> dict[str, Any]:
        with self.lock:
            self._discard_old_previews()
            context = self.course_contexts.get(token)
            if context is None:
                raise TimetableError(
                    "時間割から引き継いだ登録情報の有効期限が切れました。",
                    code="course_context_expired", status=404,
                )
            self.get_preview(context.preview_token)
            preview = self.previews[context.preview_token]
            match = self._subject_usage_summary(preview, context.subject_name)
            if match is None:
                raise TimetableError(
                    "この項目は現在の時間割にありません。",
                    code="subject_context_invalid", status=409,
                )
            if _mapping_key(context.subject_name) in self._config()["subjectMappings"]:
                raise TimetableError(
                    "この科目はすでにClassViewの授業と対応済みです。",
                    code="subject_already_mapped", status=409,
                )
            return {
                "token": token,
                "previewToken": context.preview_token,
                "returnUrl": f"/timetable?preview={context.preview_token}",
                **match,
            }

    def complete_course_registration(self, token: str, course_id: str) -> dict[str, Any]:
        with self.lock:
            context_payload = self.get_course_context(token)
            mapping = self.save_mapping(
                context_payload["subjectName"], course_id, replace_existing=True
            )
            preview = self.refresh_preview(context_payload["previewToken"])
            self.course_contexts.pop(token, None)
            return {
                "mapping": mapping,
                "preview": preview,
                "returnUrl": context_payload["returnUrl"],
            }

    def cancel_course_context(self, token: str) -> dict[str, Any]:
        with self.lock:
            context = self.course_contexts.pop(token, None)
            return {
                "success": True,
                "returnUrl": f"/timetable?preview={context.preview_token}" if context else "/timetable",
            }

    def refresh_preview(self, token: str) -> dict[str, Any]:
        with self.lock:
            self._discard_old_previews()
            preview = self.previews.get(token)
            if preview is None:
                raise TimetableError(
                    "確認内容の有効期限が切れました。Excelをもう一度読み込んでください。",
                    code="preview_expired", status=404,
                )
            config = self._config()
            valid_course_ids = {item["id"] for item in self._course_options()}
            retained = [
                issue for issue in preview.issues
                if issue["code"] not in {
                    "unregistered_subject", "mapped_course_missing",
                    "unknown_group_tag", "duplicate_entry",
                }
            ]
            source_index = {
                (
                    item.get("date"), item.get("period"), item.get("subjectRaw"),
                    item.get("rawToken"),
                ): item
                for item in preview.token_occurrences
            }
            for entry in preview.entries:
                parsed = self.parse_subject(entry["subjectRaw"])
                mapping_name = parsed["mappingName"]
                resolution = self._resolve_subject_mapping(parsed, config, valid_course_ids)
                course_id = resolution["courseId"]
                if resolution["missingCourseId"]:
                    retained.append(self._issue(
                        "blocking", "mapped_course_missing",
                        f"「{mapping_name}」の対応先授業が見つかりません: {resolution['missingCourseId']}",
                        entry=entry,
                        extra={
                            "subjectName": mapping_name,
                            "courseId": resolution["missingCourseId"],
                        },
                    ))
                elif (
                    not resolution["explicitItem"]
                    and resolution["kind"] in {"class", "exam"}
                    and mapping_name
                    and course_id is None
                ):
                    retained.append(self._issue(
                        "blocking", "unregistered_subject",
                        f"未登録の科目があります: {mapping_name}", entry=entry,
                        extra={"subjectName": mapping_name},
                    ))
                resolved_groups = []
                for group_token in (
                    item for item in parsed["tokens"] if item["tokenType"] == "group"
                ):
                    alias = self._group_alias(group_token["rawToken"])
                    canonical_id = config["groupMappings"].get(alias)
                    canonical = config["canonicalGroups"].get(canonical_id, {}) if canonical_id else {}
                    display_name = _text(canonical.get("displayName")) if isinstance(canonical, dict) else ""
                    if display_name:
                        resolved_groups.append(display_name)
                        continue
                    resolved_groups.append(group_token["tokenValue"])
                    source = source_index.get((
                        entry.get("date"), entry.get("period"), entry.get("subjectRaw"),
                        group_token["rawToken"],
                    ), {})
                    retained.append(self._issue(
                        "blocking", "unknown_group_tag",
                        f"未登録の受講グループ表記があります: {group_token['rawToken']}",
                        entry=entry,
                        extra={
                            "groupTag": group_token["tokenValue"],
                            "rawGroupToken": group_token["rawToken"],
                            "groupAlias": alias,
                            "sourceRow": source.get("sourceRow"),
                            "sourceCell": source.get("sourceCell"),
                        },
                    ))
                entry["courseId"] = course_id
                entry["kind"] = resolution["kind"]
                entry["groupTags"] = list(dict.fromkeys(resolved_groups))
                entry["entryId"] = self._entry_id(entry)
            self._detect_duplicates(preview.entries, retained)
            preview.issues = self._deduplicate_issues(retained)
            preview.changes = self.compare(preview.comparison_base, preview.entries)
            preview.merged_snapshot = sorted(preview.merged_snapshot, key=self._entry_sort_key)
            preview.token_summary = self._token_summary(preview.token_occurrences, config)
            return self._preview_payload(preview)

    def save_preview(self, token: str, warnings_acknowledged: bool) -> dict[str, Any]:
        with self.lock:
            self._discard_old_previews()
            preview = self.previews.get(token)
            if preview is None:
                raise TimetableError("確認内容の有効期限が切れました。Excelをもう一度読み込んでください。", code="preview_expired", status=404)
            blocking = [issue for issue in preview.issues if issue["severity"] == "blocking"]
            warnings = [issue for issue in preview.issues if issue["severity"] == "warning"]
            if blocking:
                raise TimetableError("解決が必要な項目が残っているため保存できません。", code="blocking_issues", status=409)
            if warnings and not warnings_acknowledged:
                raise TimetableError("確認が必要な項目を確認済みにしてから保存してください。", code="warnings_not_acknowledged", status=409)
            documents = self._public_documents(preview.merged_snapshot, preview.periods)
            backup_path = self._write_public_documents(documents)
            snapshot = {
                "schemaVersion": 1,
                "savedAt": _iso_now(),
                "entries": preview.merged_snapshot,
                "periods": preview.periods,
            }
            metadata = {
                "lastImportedAt": _iso_now(),
                "sourceFilename": preview.filename,
                "sourceModifiedAt": preview.source_modified_at,
                "sheetName": preview.sheet_name,
                "confirmedStartDate": preview.start_date.isoformat(),
                "confirmedEndDate": preview.end_date.isoformat(),
            }
            config = self._config()
            config["sheetName"] = preview.sheet_name
            try:
                for private_path in (self.config_path, self.snapshot_path, self.metadata_path):
                    if private_path.exists():
                        shutil.copy2(private_path, backup_path / f"private-{private_path.name}")
                self._atomic_json(self.config_path, config)
                self._atomic_json(self.snapshot_path, snapshot)
                self._atomic_json(self.metadata_path, metadata)
            except Exception as error:
                self._restore_save_backup(backup_path)
                if isinstance(error, TimetableError):
                    raise
                raise TimetableError(
                    "時間割の保存を完了できませんでした。元のデータは復元されています。",
                    code="write_failed",
                    status=500,
                ) from error
            self.previews.pop(token, None)
            return {
                "success": True,
                "message": "時間割を保存しました。内容はまだ公開されていません。",
                "entryCount": len(preview.merged_snapshot),
                "changedFileCount": len(documents),
                "backupPath": str(backup_path),
                "unpublished": True,
            }

    def _restore_save_backup(self, backup: Path) -> None:
        self.public_root.mkdir(parents=True, exist_ok=True)
        for path in self.public_root.glob("*.json"):
            path.unlink(missing_ok=True)
        for path in backup.glob("*.json"):
            if not path.name.startswith("private-"):
                shutil.copy2(path, self.public_root / path.name)
        for private_path in (self.config_path, self.snapshot_path, self.metadata_path):
            saved = backup / f"private-{private_path.name}"
            if saved.exists():
                shutil.copy2(saved, private_path)
            else:
                private_path.unlink(missing_ok=True)

    def _public_documents(self, entries: list[dict[str, Any]], periods: dict[str, dict[str, str]]) -> dict[str, dict[str, Any]]:
        months: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
        for entry in entries:
            public_entry = {
                key: copy.deepcopy(value)
                for key, value in entry.items()
                if key != "date" and not key.startswith("_")
            }
            months[entry["date"][:7]][entry["date"]].append(public_entry)
        updated_at = _iso_now()
        version_source = {"entries": entries, "periods": periods}
        serialized_entries = json.dumps(version_source, ensure_ascii=False, sort_keys=True).encode("utf-8")
        version = hashlib.sha256(serialized_entries).hexdigest()[:12]
        academic_years = sorted({
            str(int(item["date"][:4]) if int(item["date"][5:7]) >= 4 else int(item["date"][:4]) - 1)
            for item in entries
        })
        academic_year = academic_years[0] if len(academic_years) == 1 else "-".join(academic_years)
        documents: dict[str, dict[str, Any]] = {
            "periods.json": {"schemaVersion": 1, "periods": periods},
            "manifest.json": {
                "schemaVersion": 1,
                "academicYear": academic_year,
                "updatedAt": updated_at,
                "version": version,
                "months": sorted(months),
            },
        }
        for month, days in months.items():
            documents[f"{month}.json"] = {
                "schemaVersion": 1,
                "yearMonth": month,
                "days": {day: sorted(items, key=lambda item: (item["period"], item["entryId"])) for day, items in sorted(days.items())},
            }
        return documents

    def _schemas(self) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
        return (
            self._load_json(self.month_schema_path, {}),
            self._load_json(self.manifest_schema_path, {}),
            self._load_json(self.periods_schema_path, {}),
        )

    def _validate_public_document(self, filename: str, document: dict[str, Any]) -> None:
        month_schema, manifest_schema, periods_schema = self._schemas()
        schema = manifest_schema if filename == "manifest.json" else periods_schema if filename == "periods.json" else month_schema
        Draft202012Validator.check_schema(schema)
        errors = sorted(Draft202012Validator(schema).iter_errors(document), key=lambda error: list(error.absolute_path))
        if errors:
            path = ".".join(str(item) for item in errors[0].absolute_path)
            raise TimetableError(
                f"時間割の検証に失敗しました（{filename}{':' + path if path else ''}）。",
                code="schema_error", status=500,
            )

    def _write_public_documents(self, documents: dict[str, dict[str, Any]]) -> Path:
        self.public_root.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d-%H%M%S-%f")
        backup = self.backups_root / f"public-{stamp}"
        backup.mkdir(parents=True)
        if self.public_root.exists():
            for path in self.public_root.glob("*.json"):
                shutil.copy2(path, backup / path.name)
        staged: list[tuple[Path, Path]] = []
        try:
            for filename, document in documents.items():
                self._validate_public_document(filename, document)
                descriptor, temp_name = tempfile.mkstemp(prefix=f".{Path(filename).stem}-", suffix=".tmp", dir=self.public_root)
                temp_path = Path(temp_name)
                with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
                    json.dump(document, handle, ensure_ascii=False, indent=2)
                    handle.write("\n")
                    handle.flush()
                    os.fsync(handle.fileno())
                if json.loads(temp_path.read_text(encoding="utf-8")) != document:
                    raise TimetableError("時間割一時ファイルの検証に失敗しました。", status=500)
                staged.append((self.public_root / filename, temp_path))
            for target, temporary in staged:
                os.replace(temporary, target)
            expected = set(documents)
            for path in self.public_root.glob("????-??.json"):
                if path.name not in expected:
                    path.unlink()
            return backup
        except Exception as error:
            for path in self.public_root.glob("*.json"):
                path.unlink(missing_ok=True)
            for path in backup.glob("*.json"):
                shutil.copy2(path, self.public_root / path.name)
            if isinstance(error, TimetableError):
                raise
            raise TimetableError("時間割を保存できませんでした。元の公開データは復元されています。", code="write_failed", status=500) from error
        finally:
            for _target, temporary in staged:
                temporary.unlink(missing_ok=True)

    @staticmethod
    def _atomic_json(path: Path, document: dict[str, Any]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temp_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".tmp", dir=path.parent)
        temporary = Path(temp_name)
        try:
            with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
                json.dump(document, handle, ensure_ascii=False, indent=2)
                handle.write("\n")
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)
